import openpyxl as xl
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import random

import openpyxl.utils.exceptions


def open_file():
    try:
        Tk().withdraw()
        filename = askopenfilename()
        wb = xl.load_workbook(filename)
        sheet = wb['list']
        return sheet
        # this aim at checking if the email column is empty.
        # In practice it checks if the first cell of the second column is empty.
    except FileNotFoundError:
        file_not_supported = "The file needs to be a valid Excel file"
        return file_not_supported
    except openpyxl.utils.exceptions.InvalidFileException:
        no_file = "You need to select a valid Excel file"
        return no_file


def raffle(sheet, num_winners):
    participants_no_email = []
    participants_with_email = {}
    if sheet.cell(row=1, column=2).value is None:
        i = 1
        for i in range(1, sheet.max_row + 1):  # the for loop iterates through the values of the first column
            name = sheet.cell(row=i, column=1).value
            participants_no_email.append(name)
            i += 1
        no_email_outcome = random.sample(participants_no_email, int(num_winners))
        return no_email_outcome
    else:
        # this condition evaluates if the second column is not empty,
        # in practice if the first cell of the second column is not empty.
        i = 1
        for i in range(1, sheet.max_row + 1):
            name = str(sheet.cell(row=i, column=1).value)
            email = str(sheet.cell(row=i, column=2).value)
            participants_with_email.update({name: email})
            i += 1
        win_keys = random.sample(sorted(participants_with_email.keys()), int(num_winners))
        with_email_outcome = {}
        for key in win_keys:
            with_email_outcome.update({key: participants_with_email[key]})
        return with_email_outcome


def export_file(winners):
    try:
        if type(winners) == list:
            new_wb = xl.Workbook()
            winners_sheet = new_wb['Sheet']
            line = 1
            for person in winners:
                winners_sheet.cell(row=line, column=1, value=person)
                line += 1
            new_wb.save(filename='winners.xlsx')
        else:
            new_wb = xl.Workbook()
            winners_sheet = new_wb['Sheet']
            line = 1
            for key in winners:
                winners_sheet.cell(row=line, column=1, value=key)
                winners_sheet.cell(row=line, column=2, value=winners[key])
                line += 1
            new_wb.save(filename='winners.xlsx')
    except TypeError:
        error_msg = "Please click Raffle before trying to export the results"
        return error_msg
